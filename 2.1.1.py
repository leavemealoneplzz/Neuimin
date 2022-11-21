import csv
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Border, Side
from openpyxl import Workbook

class Vacancies:
    currency = {
        "AZN": 35.68, "BYR": 23.91, "EUR": 59.90, "GEL": 21.74, "KGS": 0.76,
        "KZT": 0.13, "RUR": 1, "UAH": 1.64, "USD": 60.66, "UZS": 0.0055,
    }
    def __init__(self, vacancy):
        self.name = vacancy['name']
        self.salary_from = int(float(vacancy['salary_from']))
        self.salary_to = int(float(vacancy['salary_to']))
        self.salary_currency = vacancy['salary_currency']
        self.salary_average = self.currency[self.salary_currency] * (self.salary_from + self.salary_to) / 2
        self.area_name = vacancy['area_name']
        self.year = int(vacancy['published_at'][:4])

class Data:
    def __init__(self, file, vacancy):
        self.file = file
        self.vacancy = vacancy

    @staticmethod
    def incremention(dic, key, amount):
        if key in dic:
            dic[key] += amount
        else:
            dic[key] = amount

    @staticmethod
    def avg(dic):
        new_dic = {}
        for key, values in dic.items():
            new_dic[key] = int(sum(values) / len(values))
        return new_dic

    @staticmethod
    def stats_print(stats1, stats2, stats3, stats4, stats5, stats6):
        print('Динамика уровня зарплат по годам: {0}'.format(stats1))
        print('Динамика количества вакансий по годам: {0}'.format(stats2))
        print('Динамика уровня зарплат по годам для выбранной профессии: {0}'.format(stats3))
        print('Динамика количества вакансий по годам для выбранной профессии: {0}'.format(stats4))
        print('Уровень зарплат по городам (в порядке убывания): {0}'.format(stats5))
        print('Доля вакансий по городам (в порядке убывания): {0}'.format(stats6))

    def stats_gen(self, count_vac, salary, salary_city, salary_vac):
        stats = self.avg(salary)
        stats2 = self.avg(salary_vac)
        stats3 = self.avg(salary_city)
        stats4 = {}
        for year, salaries in salary_city.items():
            stats4[year] = round(len(salaries) / count_vac, 4)
        stats4 = list(filter(lambda a: a[-1] >= 0.01, [(key, value) for key, value in stats4.items()]))
        stats4.sort(key=lambda a: a[-1], reverse=True)
        stats5 = stats4.copy()
        stats4 = dict(stats4)
        stats3 = list(filter(lambda a: a[0] in list(stats4.keys()), [(key, value) for key, value in stats3.items()]))
        stats3.sort(key=lambda a: a[-1], reverse=True)
        stats3 = dict(stats3[:10])
        stats5 = dict(stats5[:10])
        return stats, stats2, stats3, stats5

    def vac_increment(self, count_vac, salary, salary_city, salary_vac):
        for vac_dict in self.csv_r():
            vac = Vacancies(vac_dict)
            self.incremention(salary, vac.year, [vac.salary_average])
            if vac.name.find(self.vacancy) != -1:
                self.incremention(salary_vac, vac.year, [vac.salary_average])
            self.incremention(salary_city, vac.area_name, [vac.salary_average])
            count_vac += 1
        return count_vac

    def csv_r(self):
        with open(self.file, mode='r', encoding='utf-8-sig') as file:
            read = csv.reader(file)
            header = next(read)
            header_len = len(header)
            for row in read:
                if '' not in row and len(row) == header_len:
                    yield dict(zip(header, row))

    def statistic(self):
        salary = {}
        salary_vac = {}
        salary_city = {}
        count_vac = 0

        count_vac = self.vac_increment(count_vac, salary, salary_city, salary_vac)

        vac_num = dict([(key, len(value)) for key, value in salary.items()])
        vac_num_name = dict([(key, len(value)) for key, value in salary_vac.items()])

        if not salary_vac:
            salary_vac = dict([(key, [0]) for key, value in salary.items()])
            vac_num_name = dict([(key, 0) for key, value in vac_num.items()])

        stats, stats2, stats3, stats5 = self.stats_gen(count_vac, salary, salary_city, salary_vac)

        return stats, vac_num, stats2, vac_num_name, stats3, stats5

class Input:
    def __init__(self):
        self.file_name = input('Введите название файла: ')
        self.vacancy_name = input('Введите название профессии: ')

        dataset = Data(self.file_name, self.vacancy_name)
        stats1, stats2, stats3, stats4, stats5, stats6 = dataset.statistic()
        dataset.stats_print(stats1, stats2, stats3, stats4, stats5, stats6)

        report = Report(self.vacancy_name, stats1, stats2, stats3, stats4, stats5, stats6)
        report.excel_gen()

class Report:
    def __init__(self, vacancy_name, stats1, stats2, stats3, stats4, stats5, stats6):
        self.wb = Workbook()
        self.vacancy_name = vacancy_name
        self.stats1 = stats1
        self.stats2 = stats2
        self.stats3 = stats3
        self.stats4 = stats4
        self.stats5 = stats5
        self.stats6 = stats6

    def excel_gen(self):
        ws = self.wb.active
        ws.title = 'Статистика по годам'
        ws.append(['Год', 'Средняя зарплата', 'Средняя зарплата - ' + self.vacancy_name, 'Количество вакансий', 'Количество вакансий - ' + self.vacancy_name])
        for year in self.stats1.keys():
            ws.append([year, self.stats1[year], self.stats3[year], self.stats2[year], self.stats4[year]])

        col_width = self.width_settings()

        for i, column_width in enumerate(col_width, 1):
            ws.column_dimensions[get_column_letter(i)].width = column_width + 2

        data = []
        data.append(['Город', 'Уровень зарплат', '', 'Город', 'Доля вакансий'])
        for (city1, value1), (city2, value2) in zip(self.stats5.items(), self.stats6.items()):
            data.append([city1, value1, '', city2, value2])
        ws2 = self.wb.create_sheet('Статистика по городам')
        for row in data:
            ws2.append(row)
        col_width = []
        for row in data:
            for i, cell in enumerate(row):
                cell = str(cell)
                if len(col_width) > i:
                    if len(cell) > col_width[i]:
                        col_width[i] = len(cell)
                else:
                    col_width += [len(cell)]

        for i, column_width in enumerate(col_width, 1):  # ,1 to start at 1
            ws2.column_dimensions[get_column_letter(i)].width = column_width + 2

        fontb = Font(bold=True)
        for col in 'ABCDE':
            ws[col + '1'].font = fontb
            ws2[col + '1'].font = fontb

        for index, _ in enumerate(self.stats5):
            ws2['E' + str(index + 2)].number_format = '0.00%'

        thin = Side(border_style='thin', color='00000000')

        for row in range(len(data)):
            for col in 'ABDE':
                ws2[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

        self.stats1[1] = 1
        for row, _ in enumerate(self.stats1):
            for col in 'ABCDE':
                ws[col + str(row + 1)].border = Border(left=thin, bottom=thin, right=thin, top=thin)

        self.wb.save('report.xlsx')

    def width_settings(self):
        data = [['Год ', 'Средняя зарплата ', ' Средняя зарплата - ' + self.vacancy_name, ' Количество вакансий',
                 ' Количество вакансий - ' + self.vacancy_name]]
        col_width = []
        for row in data:
            for i, cell in enumerate(row):
                if len(col_width) > i:
                    if len(cell) > col_width[i]:
                        col_width[i] = len(cell)
                else:
                    col_width += [len(cell)]
        return col_width

if __name__ == '__main__':
    Input()